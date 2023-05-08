from django.shortcuts import render
from django.db import models
from .models import *
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse,HttpResponse,HttpResponseRedirect
import json
from django.core import serializers
from django.http import HttpResponse, FileResponse
from django.shortcuts import render
from django.templatetags.static import static
from openpyxl import load_workbook
from django.conf import settings
import os
from datetime import datetime
from django.db.models import Q
from fpdf import FPDF
# Create your views here.
def login(request):
    if request.method == 'POST':
        print("post")
        username = request.POST.get('username')
        password = request.POST.get('password')
        username_ = Employee.objects.filter(NAME_USER=username)
        print(username_[0].PASSWORD,username,password)
        # try:
        if username_[0].PASSWORD == password:
           return render(request,'index.html')
        # except:
        #      context = {
        #           'data':'ng'
        #      }
        #      return HttpResponseRedirect("/",context)
        else:
            return render(request,'login.html')
    else:
	    return render(request,'login.html')

def index(request):
	return render(request,'index.html')
def Product(request):
	if request.method == 'POST':
		res = dict()
		if  request.POST.get('action') == 'edit':
			data = Product_ob.objects.filter(product_id=request.POST.get('id')).update(product_price=request.POST.get('price'),product_qty=request.POST.get('qty'))
			res['data'] = 'ok'
			return JsonResponse(res,status=200)
		elif  request.POST.get('action') == 'delete':
			data = Product_ob.objects.filter(product_id=request.POST.get('id')).delete()
			res['data'] = 'ok'
			return JsonResponse(res,status=200)
		else:
			item_number = request.POST.get('itemnumber')
			description = request.POST.get('description')
			price = request.POST.get('price')
			quantity = request.POST.get('quantity')
			new_product = Product_ob.objects.create(product_itemno=item_number, product_description=description, product_price=price, product_qty=quantity)
			res['data'] = 'ok'
			list_product = Product_ob.objects.all()
			out_ = []
			for i in list_product:
				res = {}
				res['id']= i.product_id
				res['item_number'] = i.product_itemno
				res['description'] = i.product_description
				res['price'] = i.product_price
				res['qty'] = i.product_qty
				res['total_price'] = i.product_price * i.product_qty
				res['timein'] = i.timein			
				out_.append(res)
			context = {"data" : out_}
			print(context)
			return render(request,'Product_page.html',context)
	else:
		list_product = Product_ob.objects.all()
		out_ = []

		for i in list_product:
			res = {}
			res['id']= i.product_id
			res['item_number'] = i.product_itemno
			res['description'] = i.product_description
			res['price'] = i.product_price
			res['qty'] = i.product_qty
			res['total_price'] = i.product_price * i.product_qty
			res['timein'] = i.timein
			
			out_.append(res)
		context = {"data" : out_}
		print(context)
		return render(request,'Product_page.html',context)

def Product_Count(request):
	
	return render(request,'Product_Count.html')
def Location_mn(request):
	if request.POST:
		res = dict()
		if  request.POST.get('action') == 'view_list':
			list_Location_rack = Location_rack.objects.filter(rack_name__rack_name_id= request.POST.get('id')).values()	
			res['data'] = list(list_Location_rack)
			print(res)
			return JsonResponse(res,status=200)
		elif  request.POST.get('action') == 'view_rack':
			list_location = Location_list.objects.select_related('Location_rack__rack_name').filter(Location_rack__rack_id=request.POST.get('id'))
			# list_location = Location_list.objects.select_related('Location_rack').filter(location_id = request.POST.get('id')).values()
			print(list_location)
			data = []
			for i in list_location:
				out_ = {}
				out_['id'] = i.location_id
				out_['location_name'] = i.location_name
				out_['rack_name'] = i.Location_rack.rack_name.rack_name
				out_['rack_location'] = i.Location_rack.rack_no
				data.append(out_)
			res['data'] = data
			print(res)
			return JsonResponse(res,status=200)
		elif request.POST.get('action') == 'add':
			rack = Location_rack.objects.get(rack_id=int(request.POST.get('id')))
			Location_list.objects.create(Location_rack = rack,location_name = request.POST.get('name'))
			res['data'] = 'ok'
			print(res)
			return JsonResponse(res,status=200)
		elif  request.POST.get('action') == 'delete':
			
			Location_list.objects.get(location_id=int(request.POST.get('id'))).delete()
			res['data'] = 'ok'
			print(res)
			return JsonResponse(res,status=200)
		else:
			pass
	else:
		list_location = Location_rack_name.objects.all()
		context ={
			'location':list_location

		}
		return render(request,'location_mn.html',context)
def Location_sr(request):
	if request.POST:
		res = dict()
		if  request.POST.get('action') == 'view_detail':
			data = Product_ob.objects.filter(product_itemno = request.POST.get('itemno')).values()
			product_location_data = product_location.objects.get(Product_ob__product_id = data[0]['product_id'])

			list_location = Location_list.objects.select_related('Location_rack__rack_name').get(location_id=product_location_data.Location_list.location_id)
			print(list_location)
			res['data'] = list(data)
			
			res['rackno'] = f"{list_location.Location_rack.rack_name.rack_name}{list_location.Location_rack.rack_no}"
			res['loc_name'] = list_location.location_name
			# res['location'] = 
			
			return JsonResponse(res,status=200)
		elif request.POST.get('action') == 'view_list':
			data = Location_list.objects.filter(Location_rack__rack_id = request.POST.get('id'),Is_active = 0).values()
			res['data'] = list(data)

			return JsonResponse(res,status=200)
		elif request.POST.get('action') == 'add':
			print(request.POST )
			product_id = Product_ob.objects.get(product_itemno = request.POST.get('itemno'))
			product_id.product_qty = product_id.product_qty + int(request.POST.get('qty'))
			product_id.save()
			if len(request.POST.get('location_id')) > 0 :
				Location_list_id = Location_list.objects.get(location_id= int(request.POST.get('location_id')))
				Location_list_id.Is_active = 1	
				Location_list_id.save()
				check_data = product_location.objects.filter(Product_ob__product_id = product_id.product_id)
				if len(check_data) == 0 :
					product_location.objects.create(Product_ob=product_id,Location_list = Location_list_id)

			print(product_id.product_itemno)
			res['data'] = 'ok'

			return JsonResponse(res,status=200)
		elif request.POST.get('action') == 'out':
			product = Product_ob.objects.get(product_itemno = request.POST.get('itemno'))
			product.product_qty = product.product_qty - int(request.POST.get('qty'))
			product.save()
			product_arrival.objects.create(Product_ob = product,location_arrival_no=request.POST.get('loc_no'),location_arrival_name=  request.POST.get('loc_name'),qty=request.POST.get('qty'),total_price = request.POST.get('total_price'))
			res['data'] = 'ok'
			return JsonResponse(res,status=200)
		else:
			pass
	else:
		list_location = Location_rack_name.objects.all()
		context ={
			'location':list_location

		}
		return render(request,'location_sr.html',context)
def Seller_his_page(request):
	list_report = report_po.objects.all()
	context ={
			'po':list_report

		}
	return render(request,'Seller_his_page.html',context)
def Seller_page(request):
	if request.POST:
		res = dict()
	else:
		list_customer = Custormer.objects.all()
		list_arrival = product_arrival.objects.filter(arrival_report =0)

		context ={
			'customer':list_customer,
			'list_arrival' : list_arrival

		}
		return render(request,'Seller_page.html',context)
def Seller_export(request):
	if request.POST:
		now = datetime.now()
		datetime_string = now.strftime("%d%m%Y%H%M%S")
		filename = os.path.join(settings.BASE_DIR, 'OutsouceInspection', 'static', 'Book1.xlsx')
		
		# Load the Excel file
		wb = load_workbook(filename)

		# Modify the data
		ws = wb.active
		print(request.POST)
		list_name = []
		for key, value in request.POST.items():
			if key.startswith('name'):
				list_name.append(value)
		
		print(list_name)
		cont = 1
		custormer = Custormer.objects.get(custormer_id = request.POST.get('cus_id'))
		ws['A5'] = custormer.customer_name
		ws['A7'] = custormer.customer_contact
		ws['D5'] = f"CT-{datetime_string}"
		ws['D7'] = now.strftime("%d/%m/%Y")
		report = report_po.objects.create(po_name = f"CT-{datetime_string}",Custormer=custormer)
		report_id = report.po_id
		sum_ = 0
		for i in list_name:
			data_detial = product_arrival.objects.get(arrival_id = int(i))
			data_detial.arrival_report = report_id
			data_detial.save()
			ws['A'+str(cont+8)] = data_detial.Product_ob.product_itemno
			ws['B'+str(cont+8)] = data_detial.Product_ob.product_description
			ws['C'+str(cont+8)] = data_detial.location_arrival_name
			ws['D'+str(cont+8)] = data_detial.location_arrival_no
			ws['E'+str(cont+8)] = float(data_detial.total_price)/float(data_detial.qty)
			ws['F'+str(cont+8)] = data_detial.qty
			ws['G'+str(cont+8)] = data_detial.total_price
			sum_ = sum_+data_detial.total_price
			
			if cont == len(list_name):
				ws['F'+str(cont+10)] = 'รวมทั้งหมด'
				ws['G'+str(cont+10)] = sum_
			cont = cont+1
		# ws['A1'] = 'New Value'

		response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
		response['Content-Disposition'] = 'attachment; filename="CT-'+ datetime_string +'.xlsx"'
		wb.save(response)
		return response

def Custormer_MN(request):
	if request.POST:
		res = dict()
		if request.POST.get('action') == 'delete':
			Custormer.objects.filter(custormer_id =  request.POST.get('id')).delete()
			res['data'] = 'ok'
			return JsonResponse(res,status=200)
		elif  request.POST.get('action') == 'add':
			Custormer.objects.create(customer_name =  request.POST.get('cus_name'), customer_contact = request.POST.get('cus_dis'))
			res['data'] = 'ok'
			return JsonResponse(res,status=200)
		
	else:

		list_cus = Custormer.objects.all()
		context ={
			'data':list_cus	

		}
		return render(request,'Custormer_MN.html',context)

def dowload(request):
	if request.POST:
		now = datetime.now()
		datetime_string = now.strftime("%d%m%Y%H%M%S")
		filename = os.path.join(settings.BASE_DIR, 'OutsouceInspection', 'static', 'Book1.xlsx')
		
		# Load the Excel file
		wb = load_workbook(filename)

		# Modify the data
		ws = wb.active
		print(request.POST)
		list_name = []
		for key, value in request.POST.items():
			if key.startswith('po_id'):
				list_name.append(value)
		
		print(list_name)
		po_detial = report_po.objects.get(po_id = int(list_name[0]))
		print(po_detial.po_id)
		cont = 1
		custormer = Custormer.objects.get(custormer_id = po_detial.Custormer.custormer_id)
		ws['A5'] = custormer.customer_name
		ws['A7'] = custormer.customer_contact
		ws['D5'] = po_detial.po_name
		ws['D7'] = po_detial.timeupdate.strftime("%d/%m/%Y")
		sum_ = 0
		list_pro_arr = product_arrival.objects.filter(arrival_report = int(list_name[0]))
		for i in list_pro_arr:
		
			ws['A'+str(cont+8)] = i.Product_ob.product_itemno
			ws['B'+str(cont+8)] = i.Product_ob.product_description
			ws['C'+str(cont+8)] = i.location_arrival_name
			ws['D'+str(cont+8)] = i.location_arrival_no
			ws['E'+str(cont+8)] = float(i.total_price)/float(i.qty)
			ws['F'+str(cont+8)] = i.qty
			ws['G'+str(cont+8)] = i.total_price
			sum_ = sum_+i.total_price
			
			if cont == len(list_name):
				ws['F'+str(cont+10)] = 'รวมทั้งหมด'
				ws['G'+str(cont+10)] = sum_
			cont = cont+1
		# ws['A1'] = 'New Value'

		response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
		response['Content-Disposition'] = 'attachment; filename="CT-'+ po_detial.po_name +'.xlsx"'
		wb.save(response)
		return response